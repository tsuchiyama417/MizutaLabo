from Bio import SeqIO
import gzip
import os
import numpy as np
np.set_printoptions(suppress=True)
import math
import cmath
import openpyxl
from collections import defaultdict
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
colors = ["red","coral","gold","olive","cyan","blue","purple","magenta","green"]
animal2 = ["Human","Gorilla","P.chi.","C.chi.","F.wh.","B.wh.","Rat","Mouse","Oposs."]
import omomi_keisan 

def main():
        
    animallist = []
    with open('animallist.txt') as f:
        for line in f:
            animallist.append(line.rstrip())
            
    animalname = []
    with open('animalname.txt') as f:
        for line in f:
            animalname.append(line.rstrip()) 


    # 取得する生物種と配列
    setting_min_length = False; setting_max_length = True; 
    animaldic = fileopen(animallist, setting_min_length, setting_max_length)
    
    # 三次元座標群の作成
    animalarr = zahyou(animaldic, animalname)

    # 主成分分析 v：固有ベクトル, e：固有値
    v, e = vec(animalarr, animalname)
    
    # 主成分分析して距離行列を作成
    # matrix(v, animalname)
    
    # 3次元 -> 2次元
    trans_xy = trans(animalarr, v, animalname)

    # DFT
    dft_plot = False
    power = fourier(trans_xy, animalname, dft_plot)
    
    # コサイン類似度
    cosine_similarity(power, animalname)
    
    
def fileopen(l, flag_min, flag_max):
    
    fastafile = r'C:\Users\k\Desktop\backup\uniprot_sprot.fasta.gz'

    d = defaultdict(str)
    for i in l:
        d[i] = ''
    l = set(l)
    
    with gzip.open(fastafile, 'rt') as handle:
        for record in SeqIO.parse(handle, 'fasta'):
            print(record)
            if record.id in l:
                d[record.id] = ''.join(list(record.seq))

    # 配列
    save_file('arr', d)

    # 最小の配列に揃える
    if flag_min == True:
        n = float('inf')
        for data in d:
            n = min(n, len(d[data]))
        for data in d:
            d[data] = d[data][0:n]
            
    # 最大の配列に揃える
    elif flag_max == True:
        n = float('inf') * -1
        for data in d:
            n = max(n, len(d[data]))
        for data in d:
            if len(d[data]) <= n:
                diff = n - len(d[data])
                d[data] = d[data][0:len(d[data])] + d[data][0:diff]

    # 整形後の配列
    save_file('arr_shaping', d)
    return d

def zahyou(d, name):

    rd = defaultdict(list)

    for i in d:
        x = 0; y = 0; z = 0
        rd[i].append([x, y, z])
        for ch in d[i]:
            if ch in omomi_keisan.dic:
                x += omomi_keisan.dic[ch][0]
                y += omomi_keisan.dic[ch][1]
                z += omomi_keisan.dic[ch][2]
                rd[i].append([x, y, z])
            else:
                print("存在しない文字です")
                print(omomi_keisan.dic[ch])
            # exit()
    # 三次元座標群
    save_file('arr_graph', rd)
    # グラフ
    save_file('arr_graph_pic', rd, name)
    
    return rd

def vec(arr, name):

    # 座標の抽出
    xs = []; ys = []; zs = []
    for i in arr:
        x = []; y = []; z = []
        for seq in arr[i]:
            x.append(seq[0]); y.append(seq[1]); z.append(seq[2]) 
        xs.append(x); ys.append(y); zs.append(z)
        
    # S = 分散共分散行列
    S_list = []
    for i in range(len(xs)):
        Sxx = np.var(xs[i], ddof = 1)
        Syy = np.var(ys[i], ddof = 1)
        Szz = np.var(zs[i], ddof = 1)
        Sxy = np.cov(xs[i], ys[i])
        Sxz = np.cov(xs[i], zs[i])
        Syz = np.cov(ys[i], zs[i])
        S = np.array([[Sxx,      Sxy[0][1], Sxz[0][1]],
                    [Sxy[0][1], Syy,       Syz[0][1]],
                    [Sxz[0][1], Syz[0][1], Szz]])
        S_list.append(S)
    
    # 分散共分散行列
    save_file('S', S_list, name)

    # 固有値・固有ベクトルを計算
    vector = []
    value = []
    for i in S_list:
        # ans[0]：固有値, ans[1]：固有ベクトル
        ans = np.linalg.eig(i)
        vector.append(ans[1])
        value.append(ans[0])
        
    # 符号を揃える（一番目の生物種に揃える）
    # + = True, - = False
    mask = [[True] * 3 for _ in range(3)]
    for i in range(3):
        for j in range(3):
            if vector[0][i][j] < 0:
                mask[i][j] = False

    # 順番を揃える（固有値で降順）
    for i in range(len(value)):
        f = 1
        while f == 1:
            f = 0
            for j in range(len(value[i]) - 1):
                if value[i][j+1] > value[i][j]:
                    value[i][j], value[i][j+1] = value[i][j+1], value[i][j]
                    for k in range(3):
                        vector[i][k][j], vector[i][k][j+1] = vector[i][k][j+1], vector[i][k][j]
                    f = 1
        
        for j in range(3):
            for k in range(3):
                if mask[j][k] == True and vector[i][j][k] < 0:
                    vector[i][j][k] *= -1
                elif mask[j][k] == False and vector[i][j][k] > 0:
                    vector[i][j][k] *= -1

    # 固有値・固有ベクトル
    save_file('ve', vector, value, name)
    # 固有値・固有ベクトルのプロット
    save_file('v_plot', vector, name)

    return vector, value

def matrix(vector, name):
    
    n = len(vector)
    
    # cos類似度を計算
    cossita = []
    for k in range(n):
        for l in range(n):
            cossita.append((vector[k][0][0]*vector[l][0][0] + vector[k][1][0]*vector[l][1][0] + vector[k][2][0]*vector[l][2][0]) 
            / (math.sqrt(vector[k][0][0]*vector[k][0][0] + vector[k][1][0]*vector[k][1][0] + vector[k][2][0]*vector[k][2][0])
             * math.sqrt(vector[l][0][0]*vector[l][0][0] + vector[l][1][0]*vector[l][1][0] + vector[l][2][0]*vector[l][2][0])))
    
    # arccos距離行列を計算
    arcsita = []
    for i in range(len(cossita)):
        if cossita[i] > 1:
            cossita[i] = 1
        arcsita.append(math.acos(cossita[i]))

    # print(arcsita) 
    # 既存のdistance_arccos.xlsxを削除して新たに作成
    path = 'distance_arccos.xlsx'
    if os.path.isfile(path):
        os.remove(path)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    wb.save('distance_arccos.xlsx')

    # ブックを取得
    book = openpyxl.load_workbook('distance_arccos.xlsx')
    # シートを取得
    sheet = book['Sheet1']

    for i in range(n):
        # セルへリストを書き込む
        sheet.cell(row=i+2,column=1).value = name[i]
        for j in range(n):
            if i == 0:
                sheet.cell(row=1,column=j+2).value = name[j]

    #セルへ結果を書き込む
    c1 = 0
    c2 = 0
    for i in range(n ** 2):
        if c1 == c2 and 1:
            sheet.cell(row=c1+2,column=c2+2).value = 0.00000000
        else:
            sheet.cell(row=c1+2,column=c2+2).value = arcsita[i]
        c2 += 1
        if (c2+2) == 11:
            c1 += 1
            c2 = 0

    #保存する
    book.save('distance_arccos.xlsx')
    
    animal2 = ["Human","Gorilla","P.chi.","C.chi.","F.wh.","B.wh.","Rat","Mouse","Oposs."]
    # 既存のd.txtを削除して新たに作成
    path = 'd.txt'
    if os.path.isfile(path):
        os.remove(path)

    f = open(path, 'w')
    f.write('    ' + str(n) + ' ')
    for i in range(n):
        f.write(animal2[i])
        f.write(' ')
    f.write('\n')
    
    c = 0
    for i in range(n):
        f.write(animal2[i])
        f.write(' ')
        
        c2 = 0
        for j in range(c, n ** 2):
            f.write(str('{:.08f}'.format(arcsita[j])))
            f.write('  ')
            c2 += 1
            if c2 == n:
                f.write('\n')
                c += n
                break
    f.close()

def trans(arr, vector, name):
    
    # 座標の抽出
    xs = []; ys = []; zs = []
    for i in arr:
        x = []; y = []; z = []
        for seq in arr[i]:
            x.append(seq[0]); y.append(seq[1]); z.append(seq[2]) 
        xs.append(x); ys.append(y); zs.append(z)

    trans_x = []; trans_y = []
    # 3次元の点を2次元の点へ変換
    for i in range(len(xs)):
        x = []; y = []
        for j in range(len(xs[i])):
            x.append(vector[i][0][0] * xs[i][j] + vector[i][1][0] * ys[i][j] + vector[i][2][0] * zs[i][j]) 
            y.append(vector[i][0][1] * xs[i][j] + vector[i][1][1] * ys[i][j] + vector[i][2][1] * zs[i][j])
        trans_x.append(x)
        trans_y.append(y)

    save_file('trans', trans_x, trans_y, name)

    l = []
    l.append(trans_x); l.append(trans_y)        
    return l

def dft(f):
    N = len(f)
    Y = []
    for t in range(N):
        y = 0j  
        for x in range(N):
            a = 2 * cmath.pi * t * x / N
            y += f[x] * cmath.e ** (-1j * a)
        Y.append(y)
    return Y

def fourier(l, name, flag):

    power = []
    for i in range(len(l[0])):
        # l[0]：第一主成分, l[1]:第二主成分
        fy = dft(l[1][i])
        # パワースペクトルを追加
        power.append(np.abs(fy))
        # 0番目は平均なのでのぞく
        power[i] = power[i][1:]
    
    
    # if flag == True:
    #     # 色
    #     colors = ["red","coral","gold","olive","cyan","blue","purple","magenta","green"]
        
    #     for i in range(len(power)):
    #         # 全体のグラフを作成
    #         Figure = plt.figure()
    #         # loc
    #         ax0 = Figure.add_subplot(1,1,1)
    #         # title
    #         ax0.set_title('DFT ' + name[i], fontname="MS Gothic")
    #         # xlabel
    #         ax0.set_xlabel('周波数[Hz]', fontname="MS Gothic")
    #         # ylabel
    #         ax0.set_ylabel('パワースペクトル', fontname="MS Gothic")
    #         # x
    #         ax0.set_xlim(-5,len(l[1][i])+5)
    #         ax0.set_xticks(list(range(0,len(l[1][i])+5,100)))
            
    #         # plot
    #         ax0.plot(power[i], color = colors[i])
            
    #         # save
    #         dir = 'dft/'
    #         plt.savefig(dir + 'dft_' + name[i]+'.png')
            
    return power

def cosine_similarity(p, name):

    save_file('power', p, name)

    # 使用する配列は半分まで
    for i in range(len(p)):
        p[i] = p[i][0:len(p[i])//2]
    
    # ベクトルの正規化
    # v：各生物種のベクトルの大きさ
    for i in range(len(p)):
        v = 0
        for j in range(len(p[i])):
            v += p[i][j] ** 2
        v = math.sqrt(v)
        for j in range(len(p[i])):
            p[i][j] = p[i][j] / v

    # アークコサイン距離行列の作成
    arccos_sita = [[0] * len(p) for _ in range(len(p))]
    for i in range(len(p) - 1): 
        for j in range(i + 1, len(p)):
            bunsi = 0
            for k in range(len(p[i])):
                bunsi += p[i][k] * p[j][k]
            # アークコサイン距離に変換
            arccos_sita[i][j] = arccos_sita[j][i] = (math.acos(bunsi))

    save_file('distance_matrix', arccos_sita)

def save_file(s, *val):

    # 生物種配列
    if s == 'arr':
        data = val[0]
        path = 'data/arr.txt'
        if os.path.isfile(path):
            os.remove(path)
        f = open(path, 'w')
        for i in data:
            f.write(i)
            f.write('\n')
            f.write('length:' + str(len(data[i])))
            f.write('\n')
            f.write(data[i])
            f.write('\n')
        f.close()
        
    # 整形後の生物種配列
    elif s == 'arr_shaping':
        data = val[0]
        path = 'data/arr_shaping.txt'
        if os.path.isfile(path):
            os.remove(path)
        f = open(path, 'w')
        for i in data:
            f.write(i)
            f.write('\n')
            f.write('length:' + str(len(data[i])))
            f.write('\n')
            f.write(data[i])
            f.write('\n')
        f.close()
        
    # 三次元座標群
    elif s == 'arr_graph':
        data = val[0]
        path = 'data/arr_graph.txt'
        if os.path.isfile(path):
            os.remove(path)
        f = open(path, 'w')
        for i in data:
            f.write(i)
            f.write('\n')
            f.write('length:' + str(len(data[i])))
            f.write('\n')
            for j in data[i]:
                for k in j:
                    f.write(str(k) + ', ')
                f.write('\n')
            f.write('\n')
        f.close()
        
    # 三次元座標群グラフ
    elif s == 'arr_graph_pic':
        data = val[0]
        name = val[1]
        path = 'data/arr_graph.png'
        if os.path.isfile(path):
            os.remove(path)
        # グラフの枠を作成
        fig = plt.figure()
        ax = fig.add_subplot(111, projection='3d')

        # X,Y,Z軸にラベルを設定
        ax.set_xlabel("x")
        ax.set_ylabel("y")
        ax.set_zlabel("z")

        #色の取得
        
        
        # 座標の抽出
        xs = []; ys = []; zs = []
        for i in data:
            x = []; y = []; z = []
            for seq in data[i]:
                x.append(seq[0]); y.append(seq[1]); z.append(seq[2]) 
            xs.append(x); ys.append(y); zs.append(z)
            
        # .plotで描画
        for i in range(len(xs)):
            plt.plot(xs[i], ys[i], zs[i], color=colors[i], label = name[i])

        #labelを表示
        plt.legend()
        # 最後に.show()を書いてグラフ表示
        # 画像を保存
        fig.savefig(path, transparent=False)
    
    # 分散共分散行列
    elif s == 'S':
        data = val[0]
        name = val[1]
        path = 'data/v_matrix.txt'
        if os.path.isfile(path):
            os.remove(path)
        f = open(path, 'w')
        for i in range(len(name)):
            f.write(name[i])
            f.write('\n')
            for j in range(len(data[i])):
                f.write(str(data[i][j]))
                f.write('\n')
            f.write('\n')
        f.close()
    
    # 固有値・固有ベクトル
    elif s == 've':
        vec = val[0]
        value = val[1]
        name = val[2]
        path = 'data/vec_val.txt'
        if os.path.isfile(path):
            os.remove(path)
        f = open(path, 'w')
        for i in range(len(name)):
            f.write(name[i])
            f.write('\n')
            f.write('value:' + str(value[i]))
            f.write('\n')
            f.write('vector:' + str(vec[i]))
            f.write('\n')
            f.write('\n')
        f.close()
        
    elif s == 'v_plot':
        vector = val[0]
        name = val[1]
        path = 'data/v_plot.png'
        if os.path.isfile(path):
            os.remove(path)
        # グラフの枠を作成
        fig = plt.figure()
        ax = fig.add_subplot(111, projection='3d')

        # X,Y,Z軸にラベルを設定
        ax.set_xlabel("x")
        ax.set_ylabel("y")
        ax.set_zlabel("z")

        # ベクトルのプロット
        for i in range(len(vector)):
            plt.plot([0, vector[i][0][0]], [0, vector[i][1][0]], [0, vector[i][2][0]], "o-", color=colors[i], label = name[i])
        # labelを表示
        plt.legend()
        fig.savefig(path, transparent=False)
        
    # 3次元 -> 2次元プロット
    elif s == 'trans':
        trans_x = val[0]
        trans_y = val[1]
        name = val[2]
        for i in range(len(trans_x)):
            plt.figure()
            path = 'data/trans/trans_' + name[i] + '.png'
            if os.path.isfile(path):
                os.remove(path)
            for j in range(len(trans_x[i])):
                plt.scatter(trans_x[i][j], trans_y[i][j], color = colors[i])
            # 画像を保存
            plt.savefig(path, transparent=False)
    
    # パワースペクトル
    elif s == 'power':
        p = val[0]
        name = val[1]
        for i in range(len(p)):
            path = 'data/power/power_' + name[i] + '.png'
            if os.path.isfile(path):
                os.remove(path)
            plt.figure()
            plt.plot(p[i], color=colors[i], label = name[i])
            #labelを表示
            plt.legend()
            # 画像を保存
            plt.savefig(path,transparent=False)
    
    # アークコサイン距離行列
    elif s == 'distance_matrix':
        arccos_sita = val[0]
        n = len(arccos_sita)
        # 既存のd_fourier.txtを削除して新たに作成
        path = 'data/d.txt'
        if os.path.isfile(path):
            os.remove(path)
        f = open(path, 'w')
        f.write('    ' + str(n) + '\n')
        for i in range(n):
            f.write(animal2[i])
            for j in range(12 - len(animal2[i])):
                f.write(' ')
            for j in range(n):
                f.write(str('{:.08f}'.format(arccos_sita[i][j])))
                f.write('  ')
            f.write('\n')
        f.close()


if __name__ == "__main__":
    main()